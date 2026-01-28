import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

# Metadaten laden
try:
    with open("transcripts/_test_metadata.json", "r", encoding="utf-8") as f:
        metadata = json.load(f)
except FileNotFoundError:
    print("FEHLER: 'transcripts/_test_metadata.json' nicht gefunden.")
    print("Bitte führe erst 'create_test_transcripts.py' aus!")
    exit()

# --- Hilfsfunktionen für Text-Extraktion ---
def extract_text_from_json(obj):
    """Extrahiert Text rekursiv aus JSON"""
    result = ""
    if isinstance(obj, dict):
        for key in ["text", "full_transcript"]:
            if key in obj and isinstance(obj[key], str):
                return obj[key]
        for value in obj.values():
            if isinstance(value, (dict, list)):
                res = extract_text_from_json(value)
                if res: return res
    elif isinstance(obj, list):
        for item in obj:
            res = extract_text_from_json(item)
            if res: return res
    return result

def find_utterances_text(obj):
    """Kombiniert Utterances falls kein Full-Transcript da ist"""
    if isinstance(obj, dict):
        if "utterances" in obj and isinstance(obj["utterances"], list):
            texts = [u.get("text", "") for u in obj["utterances"] if isinstance(u, dict)]
            return " ".join(texts)
        for value in obj.values():
            res = find_utterances_text(value)
            if res: return res
    elif isinstance(obj, list):
        for item in obj:
            res = find_utterances_text(item)
            if res: return res
    return ""

def identify_structure(content):
    """Erkennt die JSON-Struktur für die Analyse"""
    if isinstance(content, dict):
        if "text" in content and "status" in content: return "Simple (text/status)"
        if "full_transcript" in content: return "Flat (full_transcript)"
        if "transcription" in content:
            t = content["transcription"]
            if "full_transcript" in t: return "Standard Gladia"
            if "utterances" in t: return "Utterances only"
        if "result" in content: return "Nested (result)"
    return "Unknown/Txt"

# --- Daten sammeln ---
test_details = []

# 1. JSON Dateien aus Metadaten
for test_info in metadata["test_cases"]:
    filename = test_info["filename"]
    try:
        with open(f"transcripts/{filename}", "r", encoding="utf-8") as f:
            content = json.load(f)
            
        text = extract_text_from_json(content)
        if not text: text = find_utterances_text(content)
        
        test_details.append({
            "filename": filename,
            "keyword": test_info["expected_keywords"],
            "text": text,
            "structure": identify_structure(content),
            "desc": test_info["description"]
        })
    except Exception as e:
        print(f"Fehler bei {filename}: {e}")

# 2. Manuell die TXT-Datei hinzufügen (damit sie nicht fehlt)
if os.path.exists("transcripts/test_apfel.txt"):
    try:
        with open("transcripts/test_apfel.txt", "r", encoding="utf-8") as f:
            text = f.read()
        test_details.append({
            "filename": "test_apfel.txt",
            "keyword": "Apfel",
            "text": text,
            "structure": "Plain Text (.txt)",
            "desc": "Manuelle Textdatei ohne JSON"
        })
    except:
        pass

# --- EXCEL ERSTELLEN ---
wb = Workbook()
ws = wb.active
ws.title = "Test Cases & Solutions"

# Styles
header_style = Font(bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="366092")
border = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))

# WICHTIG: Die neue Spaltenreihenfolge!
# Spalte A: Dateiname (damit Grader es findet)
# Spalte B: Soll-Text (damit Grader es findet)
headers = [
    "Dateiname",            # Col A -> WICHTIG FÜR GRADER
    "Soll_Text",           # Col B -> WICHTIG FÜR GRADER
    "ID",                   # Col C
    "JSON-Struktur",        # Col D
    "Extrahierter Text",    # Col E
    "Beschreibung",         # Col F
    "Ähnlichkeit (Est.)"    # Col G
]

# Header schreiben
for col, val in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=val)
    cell.font = header_style
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")
    cell.border = border

# Daten schreiben
for i, test in enumerate(test_details, 2):
    # A: Dateiname
    ws.cell(row=i, column=1, value=test["filename"]).border = border
    
    # B: Soll-Text (Keyword) - Grün markiert
    c = ws.cell(row=i, column=2, value=test["keyword"])
    c.border = border
    c.fill = PatternFill("solid", fgColor="C6EFCE")
    c.font = Font(bold=True)
    
    # C: ID
    ws.cell(row=i, column=3, value=i-1).border = border
    
    # D: Struktur - Blau
    c = ws.cell(row=i, column=4, value=test["structure"])
    c.border = border
    c.fill = PatternFill("solid", fgColor="D9E1F2")
    
    # E: Text (gekürzt)
    text_short = test["text"][:60] + "..." if len(test["text"]) > 60 else test["text"]
    ws.cell(row=i, column=5, value=text_short).border = border
    
    # F: Beschreibung
    ws.cell(row=i, column=6, value=test["desc"]).border = border

    # G: Ähnlichkeit schätzen (Visuell)
    # Simple Logik für Visualisierung
    est = "~100%"
    if "tippfehler" in test["filename"]:
        if "spielplatz" in test["filename"]: est = "~73% (Kritisch)"
        elif "apfel" in test["filename"]: est = "~91% (OK)"
        elif "bibliothek" in test["filename"]: est = "~90% (OK)"
        elif "bus" in test["filename"]: est = "~67% (Kritisch)"
    
    c = ws.cell(row=i, column=7, value=est)
    c.border = border
    if "Kritisch" in est: c.font = Font(color="FF0000")

# Spaltenbreiten
ws.column_dimensions['A'].width = 30 # Dateiname
ws.column_dimensions['B'].width = 15 # Soll
ws.column_dimensions['C'].width = 5  # ID
ws.column_dimensions['D'].width = 20 # Struktur
ws.column_dimensions['E'].width = 40 # Text
ws.column_dimensions['F'].width = 50 # Desc
ws.column_dimensions['G'].width = 15 # Est

# Speichern
wb.save("Lösungen.xlsx")
print("   Datei 'Lösungen.xlsx' im korrekten Format erstellt!")
print("   Spalte A: Dateiname")
print("   Spalte B: Soll-Text")
print("   Du kannst jetzt 'textsimilaritygrader.py' starten.")
