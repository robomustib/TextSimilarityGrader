"""
TextSimilarityGrader (https://github.com/robomustib/TextSimilarityGrader/)
Copyright (c) 2026 Mustafa Bilgin
Licensed under Creative Commons Attribution-NonCommercial 4.0 International (CC BY-NC 4.0)
"""

import os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ==========================================
# SETTINGS
# ==========================================
TRANSCRIPT_FOLDER = Path("./transcripts")
OUTPUT_FILE = "Solutions.xlsx"

def main():
    print(f"Creating Excel list from: {TRANSCRIPT_FOLDER}")

    # 1. Prepare Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Fill Solutions"

    # Colors & Styles
    header_style = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="366092") # Dark Blue
    green_fill = PatternFill("solid", fgColor="C6EFCE")  # Light Green (for Target Text)
    yellow_fill = PatternFill("solid", fgColor="FFEB9C") # Yellow (for Examples)
    border = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))

    # Set Headers
    headers = ["Filename", "Target_Text (Synonyms comma-separated)", "ID", "Type", "Status"]
    
    for col, val in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=val)
        cell.font = header_style
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    # ==========================================
    # 2. THE TWO EXAMPLES
    # ==========================================
    examples = [
        # Here we show her directly how it works:
        ("EXAMPLE_FILE_1.json", "House, Building, Hut", "Ex.", "JSON", "Example (please ignore)"),
        ("EXAMPLE_FILE_2.txt",  "Apple",                "Ex.", "TXT",  "Example (please ignore)")
    ]

    row_idx = 2
    for fname, target, id_val, type_, status in examples:
        ws.cell(row=row_idx, column=1, value=fname).border = border
        c = ws.cell(row=row_idx, column=2, value=target)
        c.border = border
        c.fill = yellow_fill
        c.font = Font(italic=True)

        ws.cell(row=row_idx, column=3, value=id_val).border = border
        ws.cell(row=row_idx, column=4, value=type_).border = border
        ws.cell(row=row_idx, column=5, value=status).border = border
        row_idx += 1

    # ==========================================
    # 3. REAL FILES
    # ==========================================
    if TRANSCRIPT_FOLDER.exists():
        # Load all files
        files = []
        for p in TRANSCRIPT_FOLDER.iterdir():
            if p.is_file() and p.suffix.lower() in ['.json', '.txt']:
                # Ignore system files starting with underscore (like _test_metadata.json)
                if not p.name.startswith("_"):
                    files.append(p)
        
        # Sort files
        files.sort(key=lambda x: x.name)

        print(f"{len(files)} files found. Inserting...")

        real_id = 1
        for p in files:
            full_filename = p.name
            ext = p.suffix.upper().replace(".", "")

            # A: Filename
            ws.cell(row=row_idx, column=1, value=full_filename).border = border
            
            # B: Target (Empty & Green)
            c = ws.cell(row=row_idx, column=2, value="")
            c.border = border
            c.fill = green_fill
            
            # C: ID
            ws.cell(row=row_idx, column=3, value=real_id).border = border
            
            # D: Type
            ws.cell(row=row_idx, column=4, value=ext).border = border
            
            # E: Status
            ws.cell(row=row_idx, column=5, value="Open").border = border

            row_idx += 1
            real_id += 1
    else:
        print("Folder 'transcripts' not found.")

    # Column Widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 25

    # Save
    try:
        wb.save(OUTPUT_FILE)
        print(f"DONE! File '{OUTPUT_FILE}' created.")
        print("The first two rows are examples.")
        print(f"Followed by {real_id-1} real files to process.")
    except PermissionError:
        print(f"Error: '{OUTPUT_FILE}' is still open. Please close it!")

if __name__ == "__main__":
    main()
