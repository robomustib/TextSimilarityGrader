"""
TextSimilarityGrader (https://github.com/robomustib/TextSimilarityGrader/)
Copyright (c) 2026 Mustafa Bilgin
Licensed under Creative Commons Attribution-NonCommercial 4.0 International (CC BY-NC 4.0)
Add-on: Blacklist Support
"""

import os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ==========================================
# SETTINGS
# ==========================================
TRANSCRIPT_FOLDER = Path("./transcripts")
OUTPUT_FILE = "Solutions_BLCK.xlsx"

def main():
    print(f"Creating Excel list from: {TRANSCRIPT_FOLDER}")

    # 1. Prepare Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Fill Solutions"

    # Colors & Styles
    header_style = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="366092") # Dark Blue
    green_fill = PatternFill("solid", fgColor="C6EFCE")  # Light Green (for Target Text)
    red_fill = PatternFill("solid", fgColor="FFC7CE")    # Light Red (for Forbidden/Blacklist)
    yellow_fill = PatternFill("solid", fgColor="FFEB9C") # Yellow (for Examples)
    border = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))

    # Set Headers (UPDATED: Added Forbidden Column)
    headers = [
        "Filename", 
        "Target_Text (Synonyms comma-separated)", 
        "Forbidden_Text (Blacklist)", 
        "ID", 
        "Type", 
        "Status"
    ]
    
    for col, val in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=val)
        cell.font = header_style
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    # ==========================================
    # 2. THE TWO EXAMPLES (UPDATED STRUCTURE)
    # ==========================================
    examples = [
        # Filename, Target, Forbidden, ID, Type, Status
        ("EXAMPLE_FILE_1.json", "House, Building", "Greenhouse, Dollhouse", "Ex.", "JSON", "Example (ignore)"),
        ("EXAMPLE_FILE_2.txt",  "Apple",           "",                      "Ex.", "TXT",  "Example (ignore)")
    ]

    row_idx = 2
    for fname, target, forbidden, id_val, type_, status in examples:
        # Col 1: Name
        ws.cell(row=row_idx, column=1, value=fname).border = border
        
        # Col 2: Target
        c = ws.cell(row=row_idx, column=2, value=target)
        c.border = border
        c.fill = yellow_fill
        c.font = Font(italic=True)

        # Col 3: Forbidden (NEW)
        c = ws.cell(row=row_idx, column=3, value=forbidden)
        c.border = border
        c.fill = red_fill
        c.font = Font(italic=True)

        # Col 4, 5, 6
        ws.cell(row=row_idx, column=4, value=id_val).border = border
        ws.cell(row=row_idx, column=5, value=type_).border = border
        ws.cell(row=row_idx, column=6, value=status).border = border
        row_idx += 1

    # ==========================================
    # 3. REAL FILES
    # ==========================================
    if TRANSCRIPT_FOLDER.exists():
        # Load all files
        files = []
        for p in TRANSCRIPT_FOLDER.iterdir():
            if p.is_file() and p.suffix.lower() in ['.json', '.txt']:
                # Ignore system files starting with underscore
                if not p.name.startswith("_"):
                    files.append(p)
        
        # Sort files
        files.sort(key=lambda x: x.name)

        print(f"{len(files)} files found. Inserting...")

        real_id = 1
        for p in files:
            full_filename = p.name
            ext = p.suffix.upper().replace(".", "")

            # A: Filename
            ws.cell(row=row_idx, column=1, value=full_filename).border = border
            
            # B: Target (Empty & Green)
            c = ws.cell(row=row_idx, column=2, value="")
            c.border = border
            c.fill = green_fill
            
            # C: Forbidden (Empty & Red - NEW)
            c = ws.cell(row=row_idx, column=3, value="")
            c.border = border
            c.fill = red_fill

            # D: ID
            ws.cell(row=row_idx, column=4, value=real_id).border = border
            
            # E: Type
            ws.cell(row=row_idx, column=5, value=ext).border = border
            
            # F: Status
            ws.cell(row=row_idx, column=6, value="Open").border = border

            row_idx += 1
            real_id += 1
    else:
        print("Folder 'transcripts' not found.")

    # Column Widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 35 # Target
    ws.column_dimensions['C'].width = 35 # Forbidden
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 25

    # Save
    try:
        wb.save(OUTPUT_FILE)
        print(f"DONE! File '{OUTPUT_FILE}' created.")
        print("Structure: Filename | Target | Forbidden | ID | Type | Status")
    except PermissionError:
        print(f"Error: '{OUTPUT_FILE}' is still open. Please close it!")

if __name__ == "__main__":
    main()