import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os
from pathlib import Path

# Load metadata
# Note: Assumes script is run from project root, so path is transcripts/...
# If running from tests/ folder, adjustment might be needed, but we assume root execution as per README.
meta_path = Path("transcripts/_test_metadata.json")

try:
    with open(meta_path, "r", encoding="utf-8") as f:
        metadata = json.load(f)
except FileNotFoundError:
    print(f"ERROR: '{meta_path}' not found.")
    print("Please run 'tests/create_test_data.py' first!")
    exit()

# --- Helper functions for text extraction ---
def extract_text_from_json(obj):
    """Recursively extracts text from JSON"""
    result = ""
    if isinstance(obj, dict):
        for key in ["text", "full_transcript"]:
            if key in obj and isinstance(obj[key], str):
                return obj[key]
        for value in obj.values():
            if isinstance(value, (dict, list)):
                res = extract_text_from_json(value)
                if res: return res
    elif isinstance(obj, list):
        for item in obj:
            res = extract_text_from_json(item)
            if res: return res
    return result

def find_utterances_text(obj):
    """Combines utterances if no full transcript is available"""
    if isinstance(obj, dict):
        if "utterances" in obj and isinstance(obj["utterances"], list):
            texts = [u.get("text", "") for u in obj["utterances"] if isinstance(u, dict)]
            return " ".join(texts)
        for value in obj.values():
            res = find_utterances_text(value)
            if res: return res
    elif isinstance(obj, list):
        for item in obj:
            res = find_utterances_text(item)
            if res: return res
    return ""

def identify_structure(content):
    """Identifies JSON structure for analysis"""
    if isinstance(content, dict):
        if "text" in content and "status" in content: return "Simple (text/status)"
        if "full_transcript" in content: return "Flat (full_transcript)"
        if "transcription" in content:
            t = content["transcription"]
            if "full_transcript" in t: return "Standard Gladia"
            if "utterances" in t: return "Utterances only"
        if "result" in content: return "Nested (result)"
    return "Unknown/Txt"

# --- Collect Data ---
test_details = []

# 1. Process JSON files from metadata
for test_info in metadata["test_cases"]:
    filename = test_info["filename"]
    try:
        with open(f"transcripts/{filename}", "r", encoding="utf-8") as f:
            content = json.load(f)
            
        text = extract_text_from_json(content)
        if not text: text = find_utterances_text(content)
        
        test_details.append({
            "filename": filename,
            "keyword": test_info["expected_keywords"],
            "text": text,
            "structure": identify_structure(content),
            "desc": test_info["description"]
        })
    except Exception as e:
        print(f"Error processing {filename}: {e}")

# 2. Manually add the plain TXT file (to ensure it's included)
txt_path = "transcripts/test_apple_plain.txt"
if os.path.exists(txt_path):
    try:
        with open(txt_path, "r", encoding="utf-8") as f:
            text = f.read()
        test_details.append({
            "filename": "test_apple_plain.txt",
            "keyword": "Apple",
            "text": text,
            "structure": "Plain Text (.txt)",
            "desc": "Manual plain text file without JSON"
        })
    except:
        pass

# --- CREATE EXCEL ---
wb = Workbook()
ws = wb.active
ws.title = "Test Cases & Solutions"

# Styles
header_style = Font(bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="366092") # Dark Blue
border = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))

# IMPORTANT: Column Layout
# Col A: Filename (for Grader)
# Col B: Target_Text (for Grader)
headers = [
    "Filename",             # Col A
    "Target_Text",          # Col B
    "ID",                   # Col C
    "JSON_Structure",       # Col D
    "Extracted_Text",       # Col E
    "Description",          # Col F
    "Similarity_Est"        # Col G
]

# Write Headers
for col, val in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=val)
    cell.font = header_style
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")
    cell.border = border

# Write Data
for i, test in enumerate(test_details, 2):
    # A: Filename
    ws.cell(row=i, column=1, value=test["filename"]).border = border
    
    # B: Target Text (Keyword) - Highlighted Green
    c = ws.cell(row=i, column=2, value=test["keyword"])
    c.border = border
    c.fill = PatternFill("solid", fgColor="C6EFCE")
    c.font = Font(bold=True)
    
    # C: ID
    ws.cell(row=i, column=3, value=i-1).border = border
    
    # D: Structure - Blueish
    c = ws.cell(row=i, column=4, value=test["structure"])
    c.border = border
    c.fill = PatternFill("solid", fgColor="D9E1F2")
    
    # E: Text (truncated)
    text_short = test["text"][:60] + "..." if len(test["text"]) > 60 else test["text"]
    ws.cell(row=i, column=5, value=text_short).border = border
    
    # F: Description
    ws.cell(row=i, column=6, value=test["desc"]).border = border

    # G: Estimated Similarity (Visual Aid)
    est = "~100%"
    fname = test["filename"]
    
    if "typo" in fname or "phonetic" in fname:
        if "bus" in fname: est = "~Check (Short Word)"
        elif "apple" in fname: est = "~91% (OK)"
        elif "library" in fname: est = "~90% (OK)"
        elif "playground" in fname: est = "~90% (OK)"
    elif "nospace" in fname:
        est = "Contains Match"
    
    c = ws.cell(row=i, column=7, value=est)
    c.border = border
    if "Check" in est: c.font = Font(color="FF0000")

# Column Widths
ws.column_dimensions['A'].width = 30 # Filename
ws.column_dimensions['B'].width = 15 # Target
ws.column_dimensions['C'].width = 5  # ID
ws.column_dimensions['D'].width = 20 # Structure
ws.column_dimensions['E'].width = 40 # Text
ws.column_dimensions['F'].width = 50 # Desc
ws.column_dimensions['G'].width = 20 # Est

# Save
output_file = "Solutions.xlsx"
wb.save(output_file)
print(f"   DONE! File '{output_file}' created successfully!")
print("   Column A: Filename")
print("   Column B: Target_Text")
print("   You can now run 'evaluate.py' to test the grading.")
